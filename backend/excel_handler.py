"""
Excel 生成处理器
负责标签替换、模板填充、Excel 导出等后端逻辑。
所有原 index.html 中的前端 JS 逻辑（getTagValue、generateSheet 等）迁移至此。
"""

import io
import json
import re
import zipfile
from datetime import datetime, date

import openpyxl
from openpyxl.utils import get_column_letter


# ============================================================
# 工具函数
# ============================================================

def safe_float(val, default=0.0):
    """安全转换数值，处理空字符串等异常情况"""
    if val is None or val == '':
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    """安全转换整数，处理空字符串等异常情况"""
    if val is None or val == '':
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def num_to_chinese_shenbao(n):
    """
    数字转人民币大写（动态自适应）
    根据实际金额大小动态决定 拾万/万/仟 等位：
      - 1253.12  → 壹仟贰佰伍拾叁元壹角贰分(¥：1253.12)
      - 11203    → 壹万壹仟贰佰零叁元零角零分(¥：11203.00)
      - 121203   → 壹拾贰万壹仟贰佰零叁元零角零分(¥：121203.00)
    对应 JS: numToChineseShenbao()
    """
    val = float(n or 0)
    int_part = int(val)  # 整数部分（元）
    dec_part = round((val - int_part) * 100)  # 小数部分（角分，2位）
    digits = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    result = ''

    if int_part >= 100000:
        # 拾万位及以上
        shiwan = int_part // 100000
        wan = (int_part % 100000) // 10000
        rest = int_part % 10000
        result += digits[shiwan] + '拾'
        if wan > 0:
            result += digits[wan] + '万'
        else:
            result += '万'
        # 仟位以下
        if rest > 0:
            q = rest // 1000
            b = (rest % 1000) // 100
            sh = (rest % 100) // 10
            g = rest % 10
            if q > 0:
                result += digits[q] + '仟'
            elif b > 0 or sh > 0 or g > 0:
                result += '零'
            if b > 0:
                result += digits[b] + '佰'
            elif q > 0 and (sh > 0 or g > 0):
                result += '零'
            if sh > 0:
                result += digits[sh] + '拾'
            elif (q > 0 or b > 0) and g > 0:
                result += '零'
            result += digits[g] + '元'
        else:
            result += '零元'
    elif int_part >= 10000:
        # 万位
        wan = int_part // 10000
        rest = int_part % 10000
        result += digits[wan] + '万'
        if rest > 0:
            q = rest // 1000
            b = (rest % 1000) // 100
            sh = (rest % 100) // 10
            g = rest % 10
            if q > 0:
                result += digits[q] + '仟'
            elif b > 0 or sh > 0 or g > 0:
                result += '零'
            if b > 0:
                result += digits[b] + '佰'
            elif q > 0 and (sh > 0 or g > 0):
                result += '零'
            if sh > 0:
                result += digits[sh] + '拾'
            elif (q > 0 or b > 0) and g > 0:
                result += '零'
            result += digits[g] + '元'
        else:
            result += '零元'
    elif int_part > 0:
        # 仟元及以下（无高位，不需要前导零）
        q = int_part // 1000
        b = (int_part % 1000) // 100
        sh = (int_part % 100) // 10
        g = int_part % 10
        if q > 0:
            result += digits[q] + '仟'
        if b > 0:
            result += digits[b] + '佰'
        elif q > 0 and (sh > 0 or g > 0):
            result += '零'
        if sh > 0:
            result += digits[sh] + '拾'
        elif b > 0 and g > 0:
            result += '零'
        if g > 0:
            result += digits[g] + '元'
        elif q > 0 or b > 0 or sh > 0:
            result += '元'
    else:
        # 0 元
        result += '零元'

    # 角分
    jiao = dec_part // 10
    fen = dec_part % 10
    result += (digits[jiao] + '角' if jiao > 0 else '零角')
    result += (digits[fen] + '分' if fen > 0 else '零分')
    result += '(¥：' + '{:.2f}'.format(val) + ')'
    return result


def get_chuxing_text(data):
    """
    获取出行方式文本
    对应 JS: getChuxingText()
    """
    items = ['飞机', '高铁/动车', '普通列车', '长途汽车', '轮船', '其他']
    keys = ['chuxing_fei', 'chuxing_gaotie', 'chuxing_putong',
            'chuxing_changtu', 'chuxing_lunchuan', 'chuxing_qita']
    parts = []
    for i, item in enumerate(items):
        checked = data.get(keys[i], 'false')
        parts.append(('☑' if checked == 'true' else '□') + item)
    return '  '.join(parts)


def get_qizhi_time_text(data):
    """
    获取起止时间文本
    对应 JS: getQizhiTimeText()
    """
    from_val = data.get('yin_qizhi_from', '')
    to_val = data.get('yin_qizhi_to', '')
    if not from_val or not to_val:
        return ''
    try:
        f = from_val.split('-')
        t = to_val.split('-')
        f_date = f'{int(f[0])}年{int(f[1])}月{int(f[2])}日'
        t_date = f'{int(t[0])}年{int(t[1])}月{int(t[2])}日'
        d1 = date(int(f[0]), int(f[1]), int(f[2]))
        d2 = date(int(t[0]), int(t[1]), int(t[2]))
        days = (d2 - d1).days + 1
        return f'{f_date}至{t_date}，共{days}天'
    except (ValueError, IndexError):
        return ''


def get_sum_time_text(data):
    """
    获取差旅费报销凭证的出差日期文本
    对应 JS: getSumTimeText()
    """
    from_val = data.get('bao_sum_from', '')
    to_val = data.get('bao_sum_to', '')
    if not from_val or not to_val:
        return ''
    try:
        f = from_val.split('-')
        t = to_val.split('-')
        f_date = f'{int(f[0])}年{int(f[1])}月{int(f[2])}日'
        t_date = f'{int(t[0])}年{int(t[1])}月{int(t[2])}日'
        d1 = date(int(f[0]), int(f[1]), int(f[2]))
        d2 = date(int(t[0]), int(t[1]), int(t[2]))
        days = (d2 - d1).days + 1
        return f'{f_date}至{t_date}，共{days}天'
    except (ValueError, IndexError):
        return ''


def get_time_from_selects(prefix, data):
    """
    从下拉选择器值获取日期文本
    对应 JS: getDateFromSelects()
    """
    year = data.get(f'{prefix}_year', '2026')
    month = data.get(f'{prefix}_month', '1')
    day = data.get(f'{prefix}_day', '1')
    return f'{int(year)}年{int(month)}月{int(day)}日'


def get_arrive_text(data):
    """
    获取到达单位+到达地点合并文本（支持多组，换行显示）
    """
    # 先获取 arriveGroupCount，如果没有则只取第一组
    count = int(data.get('arriveGroupCount', '1'))
    lines = []
    for i in range(1, count + 1):
        dw = data.get(f'yin_arrive_danwei_{i}', '')
        dd = data.get(f'yin_arrive_didian_{i}', '')
        if dw and dd:
            lines.append(f'{dw}，{dd}')
        elif dw:
            lines.append(dw)
        elif dd:
            lines.append(dd)
    return '\n'.join(lines) if lines else ''


# ============================================================
# 伙食补助费辅助函数
# ============================================================

def parse_huo_time_data(data):
    """
    解析 huoTimeData JSON 字符串为 Python 对象
    """
    raw = data.get('huoTimeData', '{}')
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return {}


def get_huo_slot_day(slot_data):
    """
    计算某时段的天数
    对应 JS: getHuoSlotDay()
    """
    if not slot_data:
        return 0
    mode = slot_data.get('mode', 'continuous')
    if mode == 'continuous':
        start = slot_data.get('start')
        end = slot_data.get('end')
        if start and end:
            try:
                sd = datetime.strptime(start, '%Y-%m-%d')
                ed = datetime.strptime(end, '%Y-%m-%d')
                days = (ed - sd).days + 1
                return max(days, 0)
            except (ValueError, TypeError):
                return 0
        return 0
    elif mode == 'individual':
        dates = slot_data.get('dates', [])
        return len(dates)
    return 0


def get_huo_slot_time_text(person_idx, slot, huo_time_data):
    """
    获取某时段的时间文本
    对应 JS: getHuoSlotTimeText()
    """
    idx_str = str(person_idx)
    td = huo_time_data.get(idx_str)
    if not td:
        td = huo_time_data.get(person_idx)  # 也支持数字key
    if not td:
        return ''
    slot_data = td.get(slot)
    if not slot_data:
        return ''
    mode = slot_data.get('mode', 'continuous')
    if mode == 'continuous':
        start = slot_data.get('start')
        end = slot_data.get('end')
        if start and end:
            try:
                sd = datetime.strptime(start, '%Y-%m-%d')
                ed = datetime.strptime(end, '%Y-%m-%d')
                return f'{sd.year % 100}年{sd.month}月{sd.day}日至{ed.year % 100}年{ed.month}月{ed.day}日'
            except (ValueError, TypeError):
                return ''
        return ''
    elif mode == 'individual':
        dates = slot_data.get('dates', [])
        return '、'.join([d.get('text', '') for d in dates])
    return ''


def get_huo_slot_shen(person_idx, slot, huo_time_data, standr):
    """
    计算某时段的申报金额
    """
    days = get_huo_slot_day_for_person(person_idx, slot, huo_time_data)
    return days * standr


def get_huo_slot_day_for_person(person_idx, slot, huo_time_data):
    """获取某个人员某时段的天数（通过 person_idx 索引）"""
    idx_str = str(person_idx)
    td = huo_time_data.get(idx_str)
    if not td:
        td = huo_time_data.get(person_idx)
    if not td:
        return 0
    slot_data = td.get(slot)
    return get_huo_slot_day(slot_data)


# ============================================================
# 标签值映射
# ============================================================

def get_tag_value(tag, data):
    """
    根据标签名从表单数据中取值
    对应 JS: getTagValue()
    """

    # ========== 因公出差审批单 tags ==========
    if tag == 'danwei':
        return data.get('yin_danwei', '')
    if tag == 'shiyou':
        return data.get('yin_shiyou', '')
    if tag == 'arrive_danwei_point':
        return get_arrive_text(data)
    if tag == 'qizhi_time':
        return get_qizhi_time_text(data)
    if tag == 'chuxing_fangshi':
        return get_chuxing_text(data)
    if tag == 'yingong_lingdao_shenpi':
        return data.get('yin_lingdao_shenpi', '')
    if tag == 'yingongchuchai_beizhu':
        return data.get('yin_beizhu', '')

    # ========== 时间 tags ==========
    if tag == 'time':
        # 根据 sheet 名称确定前缀
        sheet_name = data.get('_current_sheet', '')
        prefix_map = {
            '因公出差审批单': 'yin',
            '伙食补助费申请表': 'huo',
            '差旅费报销凭证': 'bao',
            '超标准乘坐交通工具审批单': 'chao',
        }
        prefix = prefix_map.get(sheet_name, 'yin')
        t = get_time_from_selects(prefix, data)
        return t

    if tag == 'biaoti_time':
        # 从所有人员的时间数据自动生成：时间\n（最小值\n至最大值）
        huo_time_data = parse_huo_time_data(data)
        min_date = None
        max_date = None

        def check_slot_dates(slot_data):
            nonlocal min_date, max_date
            if not slot_data:
                return
            mode = slot_data.get('mode', 'continuous')
            if mode == 'continuous':
                start = slot_data.get('start')
                end = slot_data.get('end')
                if start and end:
                    try:
                        sd = datetime.strptime(start, '%Y-%m-%d')
                        ed = datetime.strptime(end, '%Y-%m-%d')
                        if min_date is None or sd.timestamp() < min_date:
                            min_date = sd.timestamp()
                        if max_date is None or ed.timestamp() > max_date:
                            max_date = ed.timestamp()
                    except (ValueError, TypeError):
                        pass
            elif mode == 'individual':
                for d_item in slot_data.get('dates', []):
                    key = d_item.get('key', '')
                    if key:
                        try:
                            t = datetime.strptime(key, '%Y-%m-%d')
                            if min_date is None or t.timestamp() < min_date:
                                min_date = t.timestamp()
                            if max_date is None or t.timestamp() > max_date:
                                max_date = t.timestamp()
                        except (ValueError, TypeError):
                            pass

        huo_count = int(data.get('huoPersonCount', '0'))
        for i in range(1, huo_count + 1):
            idx_str = str(i)
            td = huo_time_data.get(idx_str)
            if not td:
                td = huo_time_data.get(i)
            if td:
                check_slot_dates(td.get('zhqin'))
                check_slot_dates(td.get('tuzhong'))

        if min_date is not None and max_date is not None:
            min_d = datetime.fromtimestamp(min_date)
            max_d = datetime.fromtimestamp(max_date)
            return f'时  间\n（{min_d.year}年{min_d.month}月{min_d.day}日\n至{max_d.year}年{max_d.month}月{max_d.day}日）'
        return ''

    # ========== 因公出差人员 tags (name1..6, buzhibie1..6, zhiji1..6, sui_gou1..6) ==========
    name_match = re.match(r'^name(\d+)$', tag)
    if name_match:
        idx = int(name_match.group(1))
        yin_count = int(data.get('yinPersonCount', '0'))
        if idx <= yin_count:
            return data.get(f'yin_name{idx}', '')
        return ''

    buzhibie_match = re.match(r'^buzhibie(\d+)$', tag)
    if buzhibie_match:
        idx = int(buzhibie_match.group(1))
        yin_count = int(data.get('yinPersonCount', '0'))
        if idx <= yin_count:
            return data.get(f'yin_buzhibie{idx}', '')
        return ''

    zhiji_match = re.match(r'^zhiji(\d+)$', tag)
    if zhiji_match:
        idx = int(zhiji_match.group(1))
        yin_count = int(data.get('yinPersonCount', '0'))
        if idx <= yin_count:
            return data.get(f'yin_zhiji{idx}', '')
        return ''

    sui_gou_match = re.match(r'^sui_gou(\d+)$', tag)
    if sui_gou_match:
        idx = int(sui_gou_match.group(1))
        yin_count = int(data.get('yinPersonCount', '0'))
        if idx <= yin_count:
            val = data.get(f'yin_sui_gou{idx}', '否')
            return '√' if val == '是' else ''
        return ''

    # ========== 伙食补助 tags - 每人2槽位：奇数槽=住勤，偶数槽=途中 ==========
    huo_time_match = re.match(r'^huo_time(\d+)$', tag)
    if huo_time_match:
        tag_idx = int(huo_time_match.group(1))
        person_idx = (tag_idx + 1) // 2  # ceil(tag_idx/2)
        slot = 'tuzhong' if tag_idx % 2 == 0 else 'zhqin'
        huo_count = int(data.get('huoPersonCount', '0'))
        if person_idx <= huo_count:
            htd = parse_huo_time_data(data)
            return get_huo_slot_time_text(person_idx, slot, htd)
        return ''

    huo_day_match = re.match(r'^huo_day(\d+)$', tag)
    if huo_day_match:
        tag_idx = int(huo_day_match.group(1))
        person_idx = (tag_idx + 1) // 2
        slot = 'tuzhong' if tag_idx % 2 == 0 else 'zhqin'
        huo_count = int(data.get('huoPersonCount', '0'))
        if person_idx <= huo_count:
            htd = parse_huo_time_data(data)
            days = get_huo_slot_day_for_person(person_idx, slot, htd)
            return str(days)
        return '0'

    huo_standr_match = re.match(r'^huo_standr(\d+)$', tag)
    if huo_standr_match:
        tag_idx = int(huo_standr_match.group(1))
        person_idx = (tag_idx + 1) // 2
        huo_count = int(data.get('huoPersonCount', '0'))
        if person_idx <= huo_count:
            return data.get(f'huo_standr{person_idx}', '100')
        return '100'

    huo_shen_match = re.match(r'^huo_shen(\d+)$', tag)
    if huo_shen_match:
        tag_idx = int(huo_shen_match.group(1))
        person_idx = (tag_idx + 1) // 2
        slot = 'tuzhong' if tag_idx % 2 == 0 else 'zhqin'
        huo_count = int(data.get('huoPersonCount', '0'))
        if person_idx <= huo_count:
            htd = parse_huo_time_data(data)
            days = get_huo_slot_day_for_person(person_idx, slot, htd)
            standr = safe_float(data.get(f'huo_standr{person_idx}', '100'))
            return str(int(days * standr))
        return '0'

    huo_he_match = re.match(r'^huo_he(\d+)$', tag)
    if huo_he_match:
        return ''

    huo_bei_match = re.match(r'^huo_bei(\d+)$', tag)
    if huo_bei_match:
        tag_idx = int(huo_bei_match.group(1))
        return '途中' if tag_idx % 2 == 0 else '住勤'

    if tag == 'huo_tianzhi':
        return data.get('huo_tianzhi', '')

    # ========== 超标准 tags ==========
    if tag == 'name_0':
        return ''
    if tag == 'time_to_point':
        return ''
    if tag == 'yingcheng':
        return ''
    if tag == 'shicheng':
        return ''
    if tag == 'yuanyin':
        return ''
    if tag == 'chao_lingdaoshenpi':
        return ''
    if tag == 'chao_bei':
        return ''
    if tag == 'chao_tianzhi':
        return ''

    # ========== 差旅费报销凭证 tags ==========
    if tag == 'name':
        return data.get('bao_name', '')
    if tag == 'count':
        return data.get('bao_count', '0')
    if tag == 'buzhibie':
        return data.get('bao_buzhibie', '')
    if tag == 'arrive_point':
        return data.get('bao_arrive_point', '')
    if tag == 'sum_time':
        return get_sum_time_text(data)
    if tag == 'carship_count':
        return data.get('bao_carship_count', '0')
    if tag == 'carship_value':
        return data.get('bao_carship_value', '0')
    if tag == 'airport_count':
        return data.get('bao_airport_count', '0')
    if tag == 'airport_value':
        return data.get('bao_airport_value', '0')
    if tag == 'tuigai_count':
        return data.get('bao_tuigai_count', '0')
    if tag == 'tuigai_value':
        return data.get('bao_tuigai_value', '0')
    if tag == 'carship_standard':
        return data.get('bao_carship_standard', '')
    if tag == 'airport_standard':
        return data.get('bao_airport_standard', '')
    if tag == 'tuigai_standard':
        return data.get('bao_tuigai_standard', '')
    if tag == 'count_sum':
        cc = safe_float(data.get('bao_carship_count', '0'))
        ac = safe_float(data.get('bao_airport_count', '0'))
        tc = safe_float(data.get('bao_tuigai_count', '0'))
        return str(int(cc + ac + tc))
    if tag == 'value_sum':
        cv = safe_float(data.get('bao_carship_value', '0'))
        av = safe_float(data.get('bao_airport_value', '0'))
        tv = safe_float(data.get('bao_tuigai_value', '0'))
        return str(int(cv + av + tv))
    if tag == 'standard_sum':
        return ''
    if tag == 'shinei_count':
        return data.get('bao_shinei_count', '0')
    if tag == 'shinei_sum':
        return data.get('bao_shinei_sum', '0')
    if tag == 'shinei_standard':
        return data.get('bao_shinei_standard', '')
    if tag == 'tuoyun_count':
        return data.get('bao_tuoyun_count', '0')
    if tag == 'tuoyun_sum':
        return data.get('bao_tuoyun_sum', '0')
    if tag == 'tuoyun_standard':
        return data.get('bao_tuoyun_standard', '')
    if tag == 'food_count':
        return data.get('bao_food_count', '0')
    if tag == 'food_sum':
        return data.get('bao_food_sum', '0')
    if tag == 'food_standard':
        return data.get('bao_food_standard', '')

    # 住宿费（0 值不写入 Excel）
    zhusu_keys = ['sheng', 'shi', 'xian', 'qi', 'sum']
    for zk in zhusu_keys:
        if tag == f'{zk}_p':
            v = data.get(f'bao_{zk}_p', '0')
            return '' if v == '0' or v == 0 or v == '' else v
        if tag == f'{zk}_d':
            v = data.get(f'bao_{zk}_d', '0')
            return '' if v == '0' or v == 0 or v == '' else v
        if tag == f'{zk}_s':
            v = data.get(f'bao_{zk}_s', '0')
            return '' if v == '0' or v == 0 or v == '' else v
        if tag == f'{zk}_shen':
            v = data.get(f'bao_{zk}_shen', '0')
            return '' if v == '0' or v == 0 or v == '' else v

    # 住宿费标准字段（模板中为 standar 无 d 结尾，始终为空）
    standard_map = {'sheng_standar': '', 'shi_standar': '',
                    'xian_standar': '', 'qi_standar': '',
                    'sum_standar': ''}
    if tag in standard_map:
        return ''

    # 计发伙食补助（0 值不写入 Excel）
    if tag == 'ji_p':
        v = data.get('bao_ji_p', '0')
        return '' if v == '0' or v == 0 or v == '' else v
    if tag == 'ji_d':
        v = data.get('bao_ji_d', '0')
        return '' if v == '0' or v == 0 or v == '' else v
    if tag == 'ji_s':
        v = data.get('bao_ji_s', '0')
        return '' if v == '0' or v == 0 or v == '' else v
    if tag == 'ji_shen':
        v = data.get('bao_ji_shen', '0')
        return '' if v == '0' or v == 0 or v == '' else v
    if tag == 'ji_standar':
        return ''

    # 其他
    if tag == 'yujiekuan':
        v = safe_float(data.get('bao_yujiekuan', '0'))
        return f'¥：{int(v)}'
    if tag == 'shenbao_sum':
        v = safe_float(data.get('bao_shenbao_sum', '0'))
        return num_to_chinese_shenbao(v)
    if tag == 'hezhun_sum':
        return ''
    if tag == 'caiwu_fuhe':
        return data.get('bao_caiwu_fuhe', '')
    if tag == 'caiwu_shenhe':
        return data.get('bao_caiwu_shenhe', '')
    if tag == 'yuwu_jingban':
        return data.get('bao_yuwu_jingban', '')

    return ''


# ============================================================
# 模板标签替换
# ============================================================

def replace_tags_in_cell(cell, data, sheet_name=''):
    """
    替换单元格中的 {tag} 标签
    """
    if cell.value is None or not isinstance(cell.value, str):
        return
    val = cell.value
    if '{' not in val:
        return
    tags = re.findall(r'\{([^{}]+)\}', val)
    if not tags:
        return
    # 将 sheet_name 注入 data，供 get_tag_value 中的 {time} 标签使用
    data['_current_sheet'] = sheet_name
    for tag in tags:
        tag_val = get_tag_value(tag, data)
        val = val.replace('{' + tag + '}', tag_val)

    # 【差旅费报销凭证】申报金额：括号前的汉字部分加粗
    if 'shenbao_sum' in tags:
        _set_shenbao_rich_text(cell, val)
        return

    cell.value = val
    # 包含换行符时自动设置自动换行并居中
    if '\n' in val:
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )


def _set_shenbao_rich_text(cell, val):
    """将申报金额单元格设为富文本：括号前汉字加粗，括号后数字正常。

    兼容处理：openpyxl 3.1.2 的富文本序列化存在已知缺陷（AttributeError:
    'TextBlock' object has no attribute 'name'），会导致整表导出 500。
    此处整体包在 try/except 中，遇到任何富文本异常时回退为纯字符串，
    保证导出一定可用（仅损失加粗样式）。
    """
    try:
        from openpyxl.cell.cell import CellRichText
        from openpyxl.cell.rich_text import TextBlock
        from openpyxl.cell.text import InlineFont

        # 查找第一个括号（半角或全角），取它前面的位置
        bracket_idx = None
        for ch in ['（', '(']:
            idx = val.find(ch)
            if idx > 0 and (bracket_idx is None or idx < bracket_idx):
                bracket_idx = idx

        if bracket_idx and bracket_idx > 0:
            bold_text = val[:bracket_idx]
            normal_text = val[bracket_idx:]
            font_size = cell.font.size or 11
            font_name = cell.font.name or '宋体'
            rt = CellRichText()
            rt.append(TextBlock(InlineFont(b=True, sz=font_size, rFont=font_name), bold_text))
            rt.append(TextBlock(InlineFont(b=False, sz=font_size, rFont=font_name), normal_text))
            cell.value = rt
        else:
            # 没有括号，回退到普通文本
            cell.value = val
    except Exception:
        # 富文本不支持时回退为纯字符串，确保导出不报错
        cell.value = val

    cell.alignment = openpyxl.styles.Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )


def replace_tags_in_sheet(ws, data, sheet_name=''):
    """
    遍历整个 sheet 替换所有标签
    """
    for row in ws.iter_rows():
        for cell in row:
            replace_tags_in_cell(cell, data, sheet_name)


# ============================================================
# Sheet 处理函数注册表
# ============================================================

SHEET_HANDLERS = {}


def register_handler(sheet_name):
    """
    装饰器：注册表格处理函数
    方便后续扩展新表格类型
    """
    def decorator(func):
        SHEET_HANDLERS[sheet_name] = func
        return func
    return decorator


# ============================================================
# 因公出差审批单
# ============================================================

@register_handler('因公出差审批单')
def handle_yingong(wb, ws, data):
    """
    因公出差审批单：标准标签替换 + 人员溢出分页。
    每页最多 6 人，超出部分自动创建续页。
    """
    # 先做标准标签替换（填充表头信息）
    replace_tags_in_sheet(ws, data, '因公出差审批单')

    # 人员数量
    yin_count = int(data.get('yinPersonCount', '0'))
    if yin_count <= 6:
        # 6 人以内，只填充当前 sheet，无需续页
        fill_yingong_people(ws, data, 1, yin_count)
        return

    # 第 1 页：填充前 6 人
    fill_yingong_people(ws, data, 1, 6)

    # 续页：每 6 人一页
    page = 0
    start = 7
    while start <= yin_count:
        page += 1
        end = min(start + 5, yin_count)
        new_ws = wb.copy_worksheet(ws)
        new_ws.title = f'因公出差审批单（续{page}）'
        # 清空当前页上的人员数据（清除标签残留）
        clear_yingong_person_rows(new_ws, 1, 6)
        # 填充分页数据
        fill_yingong_people(new_ws, data, start, end)
        start += 6


def fill_yingong_people(ws, data, start_idx, end_idx):
    """填充因公出差审批单的人员数据到指定行范围"""
    template_row = 8  # 模板中第1个人的行号
    for pi in range(start_idx, end_idx + 1):
        ri = template_row + (pi - start_idx)
        name_val = data.get(f'yin_name{pi}', '')
        buzhibie = data.get(f'yin_buzhibie{pi}', '')
        zhiji = data.get(f'yin_zhiji{pi}', '')
        sui_gou = data.get(f'yin_sui_gou{pi}', '否')
        ws.cell(row=ri, column=1, value=name_val)
        ws.cell(row=ri, column=2, value=buzhibie)
        ws.cell(row=ri, column=5, value=zhiji)
        ws.cell(row=ri, column=6, value=sui_gou)


def clear_yingong_person_rows(ws, start, end):
    """清空因公出差审批单指定人员范围的行"""
    template_row = 8
    for i in range(start, end + 1):
        ri = template_row + (i - 1)
        for ci in [1, 2, 5, 6]:
            ws.cell(row=ri, column=ci, value='')


# ============================================================
# 伙食补助费申请表
# ============================================================

@register_handler('伙食补助费申请表')
def handle_huo(wb, ws, data):
    """
    伙食补助费申请表特殊处理：
    每人最多2槽位（住勤+途中），
    按行分页，每页最多 12 行数据行（第4-15行），
    超出部分自动创建续页。
    自适应不同数据组合（有人只有住勤/只有途中时，一页能放更多人）。
    """
    # 第一步：先做标准标签替换（填充表头信息）
    replace_tags_in_sheet(ws, data, '伙食补助费申请表')

    # 解析 huoTimeData
    huo_time_data = parse_huo_time_data(data)
    huo_count = int(data.get('huoPersonCount', '0'))

    # 列映射：A=序号(1), B=姓名(2), C=时间(3), D=天数(4), E=标准(5), F=申报金额(6), G=核准金额(7), H=备注(8)
    # 先收集所有非空数据行
    MAX_ROWS_PER_PAGE = 12
    all_data_rows = []

    for pi in range(huo_count):
        p_idx = pi + 1
        name_val = data.get(f'huo_name{p_idx}', '')
        if not name_val:
            continue
        standr = safe_float(data.get(f'huo_standr{p_idx}', '100'))

        # 住勤
        zhqin_day = get_huo_slot_day_for_person(p_idx, 'zhqin', huo_time_data)
        zhqin_shen = zhqin_day * standr
        zhqin_time = get_huo_slot_time_text(p_idx, 'zhqin', huo_time_data)
        if zhqin_time:
            all_data_rows.append([name_val, zhqin_time, int(zhqin_day), int(standr), int(zhqin_shen), '', '住勤'])

        # 途中
        tuzhong_day = get_huo_slot_day_for_person(p_idx, 'tuzhong', huo_time_data)
        tuzhong_shen = tuzhong_day * standr
        tuzhong_time = get_huo_slot_time_text(p_idx, 'tuzhong', huo_time_data)
        if tuzhong_time:
            all_data_rows.append([name_val, tuzhong_time, int(tuzhong_day), int(standr), int(tuzhong_shen), '', '途中'])

    if not all_data_rows:
        # 没有数据行，清空数据区域
        for ri in range(4, 16):
            for ci in range(1, 9):
                ws.cell(row=ri, column=ci, value='')
        return

    # 分页写入
    total_rows = len(all_data_rows)
    for page_idx in range(0, total_rows, MAX_ROWS_PER_PAGE):
        chunk = all_data_rows[page_idx:page_idx + MAX_ROWS_PER_PAGE]

        if page_idx == 0:
            # 第 1 页：写入当前 sheet
            target_ws = ws
        else:
            # 续页：复制当前 sheet
            target_ws = wb.copy_worksheet(ws)
            page_num = page_idx // MAX_ROWS_PER_PAGE
            target_ws.title = f'伙食补助费申请表（续{page_num}）'
            # 清空数据区域
            for ri in range(4, 16):
                for ci in range(1, 9):
                    target_ws.cell(row=ri, column=ci, value='')

        # 写入数据行
        for ri, row_data in enumerate(chunk):
            target_ws.cell(row=4 + ri, column=1, value=ri + 1)  # 序号（每页从1开始）
            target_ws.cell(row=4 + ri, column=2, value=row_data[0])  # 姓名
            target_ws.cell(row=4 + ri, column=3, value=row_data[1])  # 时间
            target_ws.cell(row=4 + ri, column=4, value=row_data[2])  # 天数
            target_ws.cell(row=4 + ri, column=5, value=row_data[3])  # 标准
            target_ws.cell(row=4 + ri, column=6, value=row_data[4])  # 申报金额
            target_ws.cell(row=4 + ri, column=7, value=row_data[5])  # 核准金额
            target_ws.cell(row=4 + ri, column=8, value=row_data[6])  # 备注

        # 清空该页剩余行
        for ri in range(4 + len(chunk), 16):
            for ci in range(1, 9):
                target_ws.cell(row=ri, column=ci, value='')


# ============================================================
# 超标准乘坐交通工具审批单
# ============================================================

@register_handler('超标准乘坐交通工具审批单')
def handle_chao(wb, ws, data):
    """超标准审批单：标准标签替换即可（单张）"""
    replace_tags_in_sheet(ws, data, '超标准乘坐交通工具审批单')


def generate_chao_sheet(person_idx, data, template_path):
    """
    生成超标准审批单（逐人）
    对应 JS: generateChaoSheet()
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb['超标准乘坐交通工具审批单']

    # 逐单元格替换标签
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or not isinstance(cell.value, str):
                continue
            val = cell.value
            if '{' not in val:
                continue
            tags = re.findall(r'\{([^{}]+)\}', val)
            if not tags:
                continue
            for tag in tags:
                tag_val = ''
                if tag == 'name_0':
                    tag_val = data.get(f'chao_name{person_idx}', '')
                elif tag == 'zhiji':
                    tag_val = data.get(f'chao_zhiji{person_idx}', '')
                elif tag == 'bubie':
                    tag_val = data.get(f'chao_bubie{person_idx}', '')
                elif tag == 'time_to_point':
                    tag_val = data.get(f'chao_time_to_point{person_idx}', '')
                elif tag == 'yingcheng':
                    tag_val = data.get(f'chao_yingcheng{person_idx}', '')
                elif tag == 'shicheng':
                    tag_val = data.get(f'chao_shicheng{person_idx}', '')
                elif tag == 'yuanyin':
                    tag_val = data.get(f'chao_yuanyin{person_idx}', '')
                elif tag == 'chao_lingdaoshenpi':
                    tag_val = data.get(f'chao_lingdaoshenpi{person_idx}', '')
                elif tag == 'chao_bei':
                    tag_val = data.get(f'chao_bei{person_idx}', '')
                elif tag == 'chao_tianzhi':
                    tag_val = data.get(f'chao_tianzhi{person_idx}', '')
                    if not tag_val:
                        tag_val = data.get('huo_tianzhi', '')
                elif tag == 'time':
                    tag_val = get_time_from_selects('chao', data)
                else:
                    tag_val = get_tag_value(tag, data)
                val = val.replace('{' + tag + '}', tag_val)
            cell.value = val
            if '\n' in val:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    # 删除其他 sheet
    _remove_other_sheets(wb, '超标准乘坐交通工具审批单')

    return wb


# ============================================================
# 差旅费报销凭证
# ============================================================

@register_handler('差旅费报销凭证')
def handle_baoxiao(wb, ws, data):
    """差旅费报销凭证：标准标签替换"""
    replace_tags_in_sheet(ws, data, '差旅费报销凭证')


# ============================================================
# Sheet 工具
# ============================================================

def _remove_other_sheets(wb, keep_name):
    """删除 workbook 中除 keep_name 及其续页外的所有 sheet"""
    to_remove = []
    for sheet in wb.worksheets:
        # 保留原 sheet 和所有以原 sheet 名开头的续页
        if not sheet.title.startswith(keep_name):
            to_remove.append(sheet.title)
    for name in to_remove:
        wb.remove(wb[name])


# ============================================================
# 主导出函数
# ============================================================

def generate_excel(selected_forms, form_data, template_path):
    """
    根据选中的表单类型，生成 Excel 文件（zip 或单个文件）

    参数:
        selected_forms: list[str] 选中的表单类型，如 ['yingong', 'huo', 'baoxiao']
        form_data: dict 所有表单字段值
        template_path: str 模板文件路径

    返回:
        tuple (bytes, str) 文件内容字节流，文件名
    """
    import io

    SHEET_MAP = {
        'yingong': ('因公出差审批单', '因公出差审批单.xlsx'),
        'huo': ('伙食补助费申请表', '伙食补助费申请表.xlsx'),
        'chao': ('超标准乘坐交通工具审批单', '超标准乘坐交通工具审批单_{name}.xlsx'),
        'baoxiao': ('差旅费报销凭证', '差旅费报销凭证.xlsx'),
    }

    # 选中多个表单 → 返回 zip
    needs_zip = len(selected_forms) > 1 or (len(selected_forms) == 1 and selected_forms[0] == 'chao' and int(form_data.get('chaoPersonCount', '1')) > 1)

    if needs_zip:
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for sf in selected_forms:
                if sf == 'chao':
                    chao_count = int(form_data.get('chaoPersonCount', '1'))
                    for pi in range(1, chao_count + 1):
                        name = form_data.get(f'chao_name{pi}', f'人员{pi}')
                        wb = generate_chao_sheet(pi, form_data, template_path)
                        buf = io.BytesIO()
                        wb.save(buf)
                        buf.seek(0)
                        zf.writestr(f'超标准乘坐交通工具审批单_{name}.xlsx', buf.getvalue())
                        wb.close()
                else:
                    sheet_name = SHEET_MAP[sf][0]
                    file_name = SHEET_MAP[sf][1]
                    wb = openpyxl.load_workbook(template_path)
                    ws = wb[sheet_name]

                    if sheet_name in SHEET_HANDLERS:
                        SHEET_HANDLERS[sheet_name](wb, ws, form_data)

                    _remove_other_sheets(wb, sheet_name)
                    buf = io.BytesIO()
                    wb.save(buf)
                    buf.seek(0)
                    zf.writestr(file_name, buf.getvalue())
                    wb.close()

        zip_buffer.seek(0)
        now = datetime.now()
        return zip_buffer.getvalue(), f'财务报销_{now.year}{now.month:02d}{now.day:02d}.zip'

    else:
            # 单个表单 → 返回单个 xlsx
            sf = selected_forms[0]
            sheet_name = SHEET_MAP[sf][0]
            file_name = SHEET_MAP[sf][1]

            if sf == 'chao':
                chao_count = int(form_data.get('chaoPersonCount', '1'))
                if chao_count > 1:
                    # 多人时仍然返回 zip
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                        for pi in range(1, chao_count + 1):
                            name = form_data.get(f'chao_name{pi}', f'人员{pi}')
                            wb = generate_chao_sheet(pi, form_data, template_path)
                            buf = io.BytesIO()
                            wb.save(buf)
                            buf.seek(0)
                            zf.writestr(f'超标准乘坐交通工具审批单_{name}.xlsx', buf.getvalue())
                            wb.close()
                    zip_buffer.seek(0)
                    now = datetime.now()
                    return zip_buffer.getvalue(), f'财务报销_{now.year}{now.month:02d}{now.day:02d}.zip'
                else:
                    wb = generate_chao_sheet(1, form_data, template_path)
                    buf = io.BytesIO()
                    wb.save(buf)
                    buf.seek(0)
                    wb.close()
                    return buf.getvalue(), file_name
            else:
                wb = openpyxl.load_workbook(template_path)
                ws = wb[sheet_name]
                if sheet_name in SHEET_HANDLERS:
                    SHEET_HANDLERS[sheet_name](wb, ws, form_data)
                _remove_other_sheets(wb, sheet_name)
                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                wb.close()
                return buf.getvalue(), file_name